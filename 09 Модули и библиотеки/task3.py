import pandas as pd
import openpyxl as op
from openpyxl.styles import Border, Side, Font
import matplotlib.pyplot as plt

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def get_years(df: pd.DataFrame):
    return pd.to_datetime(df['created_at']).dt.year


def get_average_salary(row):
    if pd.isna(row['salary_from']):
        return row['salary_to']
    if pd.isna(row['salary_to']):
        return row['salary_from']
    return (row['salary_from'] + row['salary_to']) / 2


def read_csv(filename):
    df = pd.read_csv(
        filename,
        header=None,
        names=['name', 'salary_from', 'salary_to', 'salary_currency',
               'area_name', 'created_at']
    )
    df['created_at'] = pd.to_datetime(df['created_at'], utc=True)
    df['average_salary'] = df.apply(get_average_salary, axis=1)
    return df


def get_city_shares(df: pd.DataFrame):
    total = len(df)
    city_counts = df['area_name'].value_counts()
    return (city_counts / total) * 100


def get_profession_mask(df: pd.DataFrame, profession: str):
    mask = df['name'].str.contains(profession, case=False, na=False)
    return df.loc[mask].copy()


def aggregate_by_year(df: pd.DataFrame, column: str, agg_func: str, profession: str = None):
    if profession:
        df = get_profession_mask(df, profession)
    years = get_years(df)
    subset = (
        df
        .groupby(years)[column]
        .agg(agg_func)
    )
    if agg_func == 'mean':
        subset = subset.round().astype(int)
    else:
        subset = subset.astype(int)
    dct = subset.to_dict()
    sorted_items = sorted(dct.items(), key=lambda item: item[0])
    return dict(sorted_items)


def get_sal_by_year(df: pd.DataFrame):
    return aggregate_by_year(df, 'average_salary', 'mean')


def get_vaccount_by_year(df: pd.DataFrame):
    return aggregate_by_year(df, 'name', 'count')


def get_profvaccount_by_year(df: pd.DataFrame, profession: str):
    return aggregate_by_year(df, 'name', 'count', profession)


def get_profsal_by_year(df: pd.DataFrame, profession: str):
    return aggregate_by_year(df, 'average_salary', 'mean', profession)


def get_sal_by_city(df: pd.DataFrame):
    city_shares = get_city_shares(df)
    valid_cities = city_shares[city_shares > 1].index
    subset = df[df['area_name'].isin(valid_cities)]
    subset = (
        subset
        .groupby('area_name')['average_salary']
        .mean()
        .round()
        .astype(int)
    )
    dct = subset.to_dict()
    sorted_items = sorted(
        dct.items(),
        key=lambda item: (-item[1], item[0])
    )
    return dict(sorted_items[:10])


def get_share_by_city(df: pd.DataFrame):
    city_shares = get_city_shares(df)
    city_shares = city_shares[city_shares > 1].round(2)
    sorted_items = sorted(
        city_shares.items(),
        key=lambda item: (-item[1], item[0])
    )
    top_10 = sorted_items[:10]
    return dict(top_10)


def create_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def setup_years_sheet(sheet):
    sheet.title = 'Статистика по годам'
    sheet['A1'] = 'Год'
    sheet['B1'] = 'Средняя зарплата'
    sheet['C1'] = 'Количество вакансий'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20


def setup_cities_sheet(sheet):
    sheet.title = 'Статистика по городам'
    sheet['A1'] = 'Город'
    sheet['B1'] = 'Уровень зарплат'
    sheet['D1'] = 'Город'
    sheet['E1'] = 'Доля вакансий, %'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['D1'].font = Font(bold=True)
    sheet['E1'].font = Font(bold=True)
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 2
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20


def fill_years_sheet(sheet, salary_by_years, vacancies_by_years, border):
    for index, (year, salary) in enumerate(salary_by_years.items()):
        row = index + 2
        sheet[f'A{row}'] = year
        sheet[f'B{row}'] = salary
    for index, (year, count) in enumerate(vacancies_by_years.items()):
        row = index + 2
        sheet[f'C{row}'] = count
    for row in sheet['A1:C17']:
        for cell in row:
            cell.border = border


def fill_cities_sheet(sheet, salary_by_cities, share_by_city, border):
    for index, (city, salary) in enumerate(salary_by_cities.items()):
        row = index + 2
        sheet[f'A{row}'] = city
        sheet[f'B{row}'] = salary
    for index, (city, share) in enumerate(share_by_city.items()):
        row = index + 2
        sheet[f'D{row}'] = city
        sheet[f'E{row}'] = share
    for row in sheet['A1:B11']:
        for cell in row:
            cell.border = border
    for row in sheet['D1:E11']:
        for cell in row:
            cell.border = border


def create_report(df: pd.DataFrame):
    wb = op.Workbook()
    years_sheet = wb.active
    cities_sheet = wb.create_sheet(title='Статистика по городам', index=1)
    setup_years_sheet(years_sheet)
    setup_cities_sheet(cities_sheet)
    salary_by_years = get_sal_by_year(df)
    vacancies_by_years = get_vaccount_by_year(df)
    salary_by_cities = get_sal_by_city(df)
    share_by_city = get_share_by_city(df)
    border = create_border()
    fill_years_sheet(years_sheet, salary_by_years, vacancies_by_years, border)
    fill_cities_sheet(cities_sheet, salary_by_cities, share_by_city, border)
    wb.save('student_works/report.xlsx')


def wrap_city_name(city):
    parts = city.replace('-', ' ').split()
    return "\n".join(parts)


def create_plot(df: pd.DataFrame, profession: str):
    salary_by_year = get_sal_by_year(df)
    prof_salary_by_year = get_profsal_by_year(df, profession)
    vacancies_by_year = get_vaccount_by_year(df)
    prof_vacancies_by_year = get_profvaccount_by_year(df, profession)
    years = list(salary_by_year.keys())
    x = list(range(len(years)))
    width = 0.4
    salary_by_city = get_sal_by_city(df)
    share_by_city = get_share_by_city(df)
    fig, sub = plt.subplots(2, 2)
    ax1, ax2, ax3, ax4 = sub.flatten()
    ax1.bar(
        [i - width / 2 for i in x],
        [salary_by_year[y] for y in years],
        width,
        label="средняя з/п"
    )
    ax1.bar(
        [i + width / 2 for i in x],
        [prof_salary_by_year.get(y, 0) for y in years],
        width,
        label=f"з/п {profession}"
    )
    ax1.set_xticks(x)
    ax1.set_xticklabels(years, rotation=90, fontsize=8)
    ax1.set_title("Уровень зарплат по годам", fontsize=8)
    ax1.grid(True, axis='y')
    ax1.legend(fontsize=8)
    ax2.bar(
        [i - width / 2 for i in x],
        [vacancies_by_year[y] for y in years],
        width,
        label="Количество вакансий"
    )
    ax2.bar(
        [i + width / 2 for i in x],
        [prof_vacancies_by_year.get(y, 0) for y in years],
        width,
        label=f"Количество вакансий {profession}"
    )
    ax2.set_xticks(x)
    ax2.set_xticklabels(years, rotation=90, fontsize=8)
    ax2.set_title("Количество вакансий по годам", fontsize=8)
    ax2.grid(True, axis='y')
    ax2.legend(fontsize=8)
    cities = list(salary_by_city.keys())
    salaries = list(salary_by_city.values())
    y_pos = range(len(cities))
    ax3.barh(y_pos, salaries)
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(
        [wrap_city_name(c) for c in cities],
        fontsize=6,
        horizontalalignment='right',
        verticalalignment='center'
    )
    ax3.set_title("Уровень зарплат по городам", fontsize=8)
    ax3.grid(True, axis='x')
    ax3.invert_yaxis()
    labels = list(share_by_city.keys())
    sizes = list(share_by_city.values())
    other = max(0.0, 100.0 - sum(sizes))
    if other > 0.1:
        labels.append("Другие")
        sizes.append(other)
    ax4.pie(
        sizes,
        labels=labels,
        textprops={'fontsize': 6},
        startangle=140
    )
    ax4.set_title("Доля вакансий по городам", fontsize=8)
    plt.tight_layout()
    plt.savefig("student_works/graph.png", dpi=200)
    plt.close()


def extract_table(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    return rows


def draw_centered_title(pdf, text, top, size):
    pdf.setFont("Arial", size)
    w = pdf.stringWidth(text, "Arial", size)
    pdf.drawString((A4[0] - w) / 2, top, text)


def draw_table(pdf, data, y_start):
    tbl = Table(data)
    tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arial'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    tw, th = tbl.wrapOn(pdf, 0, 0)
    tbl.drawOn(pdf, (A4[0] - tw) / 2, y_start - th)
    return y_start - th - 20


def create_pdf():
    pdfmetrics.registerFont(TTFont("Arial", "Arial.ttf"))
    c = canvas.Canvas("student_works/report.pdf", pagesize=A4)
    W, H = A4
    y = H - 50
    draw_centered_title(c, "Аналитика профессий по зарплатам и городам", y, 26)
    y -= 60
    c.drawImage("student_works/graph.png", W * 0.2, y - 220, width=W * 0.6, height=220)
    y -= 250
    book = op.load_workbook("student_works/report.xlsx")
    ws_y = book["Статистика по годам"]
    ws_c = book["Статистика по городам"]
    y = draw_table(c, extract_table(ws_y), y)
    c.showPage()
    y = H - 40
    draw_table(c, extract_table(ws_c), y)
    c.save()


def main():
    df = read_csv("vacancies.csv")
    create_report(df)
    profession = input().strip()
    create_plot(df, profession)
    create_pdf()


if __name__ == "__main__":
    main()
